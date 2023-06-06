from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, ScatterChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from typing import List
from score_structure import Student, TestInfo, Score, Subject, SUBJECT_NAME, SUBJECT_NUM, SUBJECT_SCORE_LIMIT
import os
import pickle
from win32com.client import Dispatch, CDispatch


def load_data(save_path: str) -> List[Student]:
    data: List[Student]
    with open(save_path, "rb") as f:
        data = pickle.load(f)
    return data


def save_data(save_path: str, data: List[Student]) -> None:
    with open(save_path, "wb") as f:
        pickle.dump(data, f)


def get_file_name_without_suffix(path: str) -> str:
    return os.path.splitext(os.path.basename(path))[0]


def standardize_dir(dir: str) -> str:
    if dir == '':
        return '\\'
    dir = dir.replace('/', '\\')        # 我是傻逼, 忘记了 replace() 并不会改变原字符串
    if dir[-1] != '\\':
        dir += '\\'
    return dir




# 该怎么命名.? format>? construct>?
def format_base_data_from_wookbook(path: str) -> List[Student]:
    wb = load_workbook(path)
    ws = wb['Sheet1']

    data: List[Student] = []
    test_info: TestInfo = TestInfo('base')

    for i in range(2, 100):             # 假定只有不超过 100 个学生
        val = ws['A' + str(i)].value
        if val == None:
            break;
        st = Student(val)
        st.push_test_back(test_info)
        data.append(st)

    data.sort()
    return data


def format_data_from_wookbook(path: str, geo_bio_point_scale: int = 100, test_name: str = 'auto') -> List[Student]:
    # 分制默认是 100, 那么忘记填的话, 成绩会保持原数据
    assert(geo_bio_point_scale in [50, 100])    # 地理生物分制应该也弄个枚举的
    wb = load_workbook(path)
    ws = wb['Sheet1']

    data: List[Student] = []
    test_info: TestInfo = TestInfo()
    if test_name == 'auto':
        test_info.set_name(get_file_name_without_suffix(path))
    else:
        test_info.set_name(test_name)


    subject_id: List[Subject] = [-1]
    for i in range(1, 20):             # 假定只有不超过 20 个学科
        val = ws[chr(ord('A') + i) + '1'].value
        if val == None:
            break
        if val in SUBJECT_NAME:
            subject_id.append(SUBJECT_NAME.index(val))
        else:
            subject_id.append(-1)

    for i in range(2, 100):             # 假定只有不超过 100 个学生
        if ws['A' + str(i)].value == None:
            break;
        # 假定第一列(A 列)是姓名, 第一行是标题
        score: Score = Score()
        for k in range(len(subject_id)):
            if subject_id[k] == -1:
                continue
            val: int = ws[chr(ord('A') + k) + str(i)].value
            if val == None:             # 未录入视为缺考(不过似乎和考零分区别也不大)
                val = -1
            if (subject_id[k] in [Subject.GEOGRAPHY, Subject.BIOLOGY]):
                val *= 100 / geo_bio_point_scale
            score.set_score_by_id(subject_id[k], val if val != None else 0)
        test_info.set_score(score)
        st = Student(ws['A' + str(i)].value)
        st.push_test_back(test_info)
        data.append(st)

    data.sort()
    return data

def extract_test_name_from_data(data: List[Student]) -> List[str]:
    test_name_list: List[str] = []

    # 根据 merge 里的规则,即使有学生没来考试,他的纪录中依然包含了这次成绩
    # 似乎让每个学生的信息都包含着考试名称有些不太好, 不过这个之后再改吧(应当是班级包含考试, 也包含学生)
    for id in range(data[0].get_test_num()):
        test = data[0].get_test(id)
        test_name_list.append(test.get_name())

    return test_name_list


def extract_student_name_from_data(data: List[Student]) -> List[str]:
    student_name_list: List[str] = []
    for stu in data:
        student_name_list.append(stu.get_name())
    return student_name_list


def merge(ori: List[Student], add: List[Student]) -> None:
    i: int = 0
    k: int = 0
    if add == []:
        return

    # 假定 add 里的学生姓名是 ori 里学生姓名的子集
    # 假定 ori 和 add 里的学生姓名都以同一种规则排好序了
    # add 里至少有一个学生考了这次试
    # 假定 add 里的学生只有一次考试的成绩
    empty_test_info = TestInfo(add[0].get_test(0).get_name(), Score())
    # print(f'empty test info is like:{empty_test_info}')

    while i < len(ori) and k < len(add):
        if ori[i].get_name() == add[k].get_name():
            ori[i].push_test_back(add[k].get_test(0))   # 一个让我不太能看出来的错误. 不是 add[k].get_test[0], 而是 add[k].get_test(0)
            k += 1
        else:
            ori[i].push_test_back(empty_test_info)
        i += 1
    pass


def draw_line_chart(path: str, stu: Student, sub: Subject, mask: List[int], geo_bio_point_scale: int) -> None:
    # 掩码为 1 则显示相应考试成绩, 为 0 则不显示
    if mask == []:
        for i in range(stu.get_test_num()): # 不传入 mask 则默认全显示
            mask.append(1)
    elif len(mask) < stu.get_test_num():    # 未指定的默认不显示(没错, 默认行为可能并不合理)
        for i in range(len(mask), stu.get_test_num()):
            mask.append(0)

    wb = Workbook()
    ws = wb.active
    ws['A1'] = '测试名称'
    ws['B1'] = '成绩'
    row = 1
    for i in range(1, stu.get_test_num()):  # 从 1 开始是因为第一场考试只是用于辅助记录所有人的名字的
        # print(mask[i])
        if mask[i - 1] == 0 or stu.get_test(i).get_score() == Score():   # 或者这场考试没有他的数据
            continue
        row += 1
        # print(stu.get_test(i))
        ws['A' + str(row)] = stu.get_test(i).get_name()
        sub_score: float = stu.get_test(i).get_score().get_score_by_id(sub)
        if sub in [Subject.GEOGRAPHY, Subject.BIOLOGY]:
            sub_score /= (100.0 / geo_bio_point_scale)
        ws['B' + str(row)] = sub_score

    # 以下内容几乎全由 new bing 教导而成. 他妈我要是懂英语我会是现在这个样子.?
    chart = LineChart()
    data = Reference(ws, min_col=2, min_row=1, max_row=row)
    chart.add_data(data, from_rows=False, titles_from_data=True)

    x_data = Reference(ws, min_col=1, min_row=2, max_col=1, max_row=row)
    chart.set_categories(x_data)

    chart.title = f'{stu.get_name()}{SUBJECT_NAME[sub]}成绩分析'
    chart.legend = None

    # chart.x_axis.title = "测试名称"
    # chart.y_axis.title = "成绩"

    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = SUBJECT_SCORE_LIMIT[sub]
    if sub in [Subject.GEOGRAPHY, Subject.BIOLOGY]:
        chart.y_axis.scaling.max /= 100 / geo_bio_point_scale

    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True
    chart.dLbls.position = 't'

    chart.style = 19
    ws.add_chart(chart, 'A1')
    wb.save(path)


def convert_xlsx_to_pdf(input_path: str, output_path: str, mode: str = 'ms') -> None:
    # 根据测试, input_path 必须是绝对路径, output_path 可以是相对路径
    print(input_path)
    print(output_path)
    dispatch_name: str = 'Excel.Application'
    if mode == 'wps':
        dispatch_name = 'Ket.Application'

    excel: CDispatch = Dispatch(dispatch_name)
    xlsx = excel.Workbooks.Open(input_path)
    xlsx.ExportAsFixedFormat(0, output_path)
    xlsx.Close()
    excel.Quit()